import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io

st.set_page_config(page_title="Maaltijden Converter", layout="centered")
st.title("🍽️ Maaltijden Converter")
st.write("Upload het ruwe Excel-bestand uit het systeem en download direct de opgemaakte versie.")

uploaded_file = st.file_uploader("Kies het ruwe Excel-bestand", type=['xlsx', 'xls'])

if uploaded_file is not None:
    if st.button("Verwerk Bestand"):
        try:
            # Lees de ruwe data direct uit de upload
            df_raw = pd.read_excel(uploaded_file, sheet_name=0, header=None)
            
            mask = df_raw.apply(lambda row: row.astype(str).str.strip().eq('Kamer').any(), axis=1)
            if not mask.any():
                st.error("Fout: Kan de kolom 'Kamer' niet vinden in dit bestand.")
                st.stop()
            
            header_row_idx = mask.idxmax()
            header_row = df_raw.iloc[header_row_idx].fillna('').astype(str).str.strip()
            
            expected_cols = ['Kamer', 'Gast(en)', 'Boeker', 'Total guests', 'Posted meals', 'Notities (gast)', 'Prijscode', 'MP Code']
            col_indices = []
            
            for expected in expected_cols:
                matches = header_row[header_row == expected]
                if not matches.empty:
                    col_indices.append(matches.index[0])
                else:
                    partial = header_row[header_row.str.contains(expected, case=False, regex=False)]
                    if not partial.empty:
                        col_indices.append(partial.index[0])
                    else:
                        st.error(f"Let op: Kolom '{expected}' niet gevonden!")
                        st.stop()

            df = df_raw.iloc[header_row_idx + 1:, col_indices].copy()
            df.columns = expected_cols
            df = df.dropna(how='all')
            
            df = df[df['Kamer'].astype(str).str.match(r'^\d+$')]
            df['Kamer'] = pd.to_numeric(df['Kamer'], errors='coerce')
            df = df.dropna(subset=['Kamer'])
            df['Kamer'] = df['Kamer'].astype(int)
            df['Total guests'] = pd.to_numeric(df['Total guests'], errors='coerce').fillna(0).astype(int)
            df['Posted meals'] = pd.to_numeric(df['Posted meals'], errors='coerce').fillna(0).astype(int)
            
            # Kamer grens verhoogd naar 6000 voor de 5900-kamers, met uitsluiting van de foute 5810
            df = df[(df['Kamer'] < 6000) & (df['Kamer'] != 5810)]
            
            display_cols = ['Kamer', 'Gast(en)', 'Boeker', 'Guests', 'Meals', 'Notities (gast)', 'Prijscode', 'MP Code']
            
            wb = Workbook()
            wb.remove(wb.active)
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            zebra_fill_1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            zebra_fill_2 = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid")
            pink_fill = PatternFill(start_color="FFD2D2", end_color="FFD2D2", fill_type="solid")
            total_font = Font(bold=True)
            total_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            def format_sheet(ws, df_subset):
                ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
                df_subset = df_subset.sort_values(by='Kamer')
                total_g = df_subset['Total guests'].sum()
                total_m = df_subset['Posted meals'].sum()
                df_display = df_subset.copy()
                df_display.columns = display_cols
                
                rows = dataframe_to_rows(df_display, index=False, header=True)
                for r_idx, row in enumerate(rows, 1):
                    ws.append(row)
                    is_group = (r_idx > 1 and "groep" in str(row[6]).lower())
                    current_fill = header_fill if r_idx == 1 else (pink_fill if is_group else (zebra_fill_1 if r_idx % 2 == 0 else zebra_fill_2))
                    for c_idx, cell in enumerate(ws[r_idx], 1):
                        cell.border = thin_border
                        cell.fill = current_fill
                        if r_idx == 1:
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        else:
                            align_kwargs = {'vertical': 'center'}
                            if c_idx == 1: align_kwargs['horizontal'] = 'left'
                            elif c_idx in [4, 5]: align_kwargs['horizontal'] = 'center'
                            elif c_idx == 6: 
                                align_kwargs['horizontal'] = 'center'
                                align_kwargs['wrap_text'] = True
                            cell.alignment = Alignment(**align_kwargs)
                
                max_row = ws.max_row + 1
                ws.cell(row=max_row, column=1, value="TOTAAL").font = total_font
                ws.cell(row=max_row, column=4, value=total_g).font = total_font
                ws.cell(row=max_row, column=5, value=total_m).font = total_font
                ws.cell(row=max_row, column=4).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=max_row, column=5).alignment = Alignment(horizontal='center', vertical='center')
                
                for c_idx in range(1, 9):
                    cell = ws.cell(row=max_row, column=c_idx)
                    if c_idx not in [1, 4, 5]: cell.value = ""
                    cell.fill = total_fill
                    cell.border = thin_border
                    
                widths = {'A': 5, 'B': 30, 'C': 30, 'D': 7, 'E': 7, 'F': 32, 'G': 30, 'H': 7}
                for col, w in widths.items():
                    ws.column_dimensions[col].width = w
                ws.column_dimensions['B'].hidden = True

            format_sheet(wb.create_sheet(title="Blok 1000 + 3000"), df[((df['Kamer'] >= 1000) & (df['Kamer'] < 2000)) | ((df['Kamer'] >= 3000) & (df['Kamer'] < 4000))])
            format_sheet(wb.create_sheet(title="Overige Kamers"), df[~(((df['Kamer'] >= 1000) & (df['Kamer'] < 2000)) | ((df['Kamer'] >= 3000) & (df['Kamer'] < 4000)))])
            
            # Sla het Excel-bestand virtueel op zodat het gedownload kan worden
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            st.success("Bestand is succesvol verwerkt!")
            
            # Download knop
            st.download_button(
                label="📥 Download Verwerkt Bestand",
                data=output,
                file_name=f"Processed_{uploaded_file.name}",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
        except Exception as e:
            st.error(f"Er ging iets mis tijdens het verwerken: {e}")
